"""Excel export — single-sheet hydraulic datasheet (openpyxl → BytesIO).

Layout (one sheet):
  Row 1–2    : Title bar
  Row 4–N    : Header block — project/line left, inlet+summary right (side-by-side)
  Row N+2+   : Transposed segment table (properties as rows, segments as columns)
"""
from __future__ import annotations

from io import BytesIO

import streamlit as st

from quickpipe.engine.results import COLUMNS

_BLUE   = "2563EB"
_NAVY   = "0F172A"
_AMBER  = "B45309"
_GREEN  = "166534"
_ALT    = "F1F5F9"
_SEC_BG = "EFF6FF"
_SEC_FG = "1E40AF"


def _styles():
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    thin = Side(style="thin",   color="CBD5E1")
    mid  = Side(style="medium", color="93C5FD")
    return {
        "title":    Font(bold=True, size=13, color=_BLUE),
        "subtitle": Font(size=9,   color="64748B"),
        "sec":      Font(bold=True, size=9,  color=_SEC_FG),
        "sec_fill": PatternFill("solid", fgColor=_SEC_BG),
        "lbl":      Font(bold=True, size=9),
        "val":      Font(size=9),
        "tbl_hdr":  Font(bold=True, color="FFFFFF", size=9),
        "tbl_fill": PatternFill("solid", fgColor=_BLUE),
        "alt_fill": PatternFill("solid", fgColor=_ALT),
        "warn":     Font(bold=True, size=9, color=_AMBER),
        "ok":       Font(bold=True, size=9, color=_GREEN),
        "bdr":      Border(left=thin, right=thin, top=thin, bottom=thin),
        "mbdr":     Border(left=mid,  right=mid,  top=mid,  bottom=mid),
        "center":   Alignment(horizontal="center", vertical="center"),
        "wrap":     Alignment(wrap_text=True, vertical="top"),
    }


def _c(ws, row, col, val="", *, font=None, fill=None, bdr=None, align=None, fmt=None):
    cell = ws.cell(row, col, val)
    if font:  cell.font           = font
    if fill:  cell.fill           = fill
    if bdr:   cell.border         = bdr
    if align: cell.alignment      = align
    if fmt:   cell.number_format  = fmt
    return cell


def _sec(ws, row, col, text, s, span=2):
    _c(ws, row, col, text, font=s["sec"], fill=s["sec_fill"], bdr=s["bdr"])
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return row + 1


def _kv(ws, row, col, label, val, s, alt=False, fmt=None):
    fill = s["alt_fill"] if alt else None
    _c(ws, row, col,     label, font=s["lbl"], fill=fill, bdr=s["bdr"])
    _c(ws, row, col + 1, val,   font=s["val"], fill=fill, bdr=s["bdr"], fmt=fmt)
    return row + 1


def build_xlsx(result, line: dict, meta: dict, solver_meta: dict) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    s  = _styles()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hydraulic datasheet"

    # Column widths — A-B = left block, C = gap, D-E = right block, F+ = segments
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 2
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 22

    tag = line.get("tag") or line.get("id", "")
    svc = line.get("service") or ""
    inlet = line.get("inlet", {})
    r0    = result.rows[0] if result.rows else None
    pipes = [x for x in result.rows if x.type == "Pipe"]
    fit_dp  = sum(x.dp_fit_kpa for x in result.rows)
    total_L = sum(x.l_m       for x in result.rows)
    max_v   = max((x.v_ms      for x in pipes), default=0.0)
    max_ve  = max((x.v_over_ve for x in pipes), default=0.0)

    # ── Row 1-2: Title bar ────────────────────────────────────────────────────
    _c(ws, 1, 1, "QUICKPIPE — HYDRAULIC LINE SIZING", font=s["title"])
    ws.merge_cells("A1:E1")
    tagline = tag + (f"  ·  {svc}" if svc else "")
    _c(ws, 2, 1, tagline, font=Font(bold=True, size=10, color=_NAVY))
    ws.merge_cells("A2:E2")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 14

    # ── Rows 4+: Left block (project + line) and right block (inlet + summary) ─
    L = 4   # left-block current row
    R = 4   # right-block current row

    # Left: Project
    L = _sec(ws, L, 1, "PROJECT", s)
    for i, (lbl, val) in enumerate([
        ("Project name",  meta.get("project_name", "")),
        ("Client",        meta.get("client", "")),
        ("Calc by",       meta.get("calc_by", "")),
        ("Checked by",    meta.get("checked_by", "")),
        ("Revision",      meta.get("rev", "")),
        ("Date",          meta.get("date", "")),
    ]):
        L = _kv(ws, L, 1, lbl, val, s, alt=bool(i % 2))
    L += 1
    L = _sec(ws, L, 1, "LINE IDENTIFICATION", s)
    L = _kv(ws, L, 1, "Tag",     tag, s)
    L = _kv(ws, L, 1, "Service", svc, s, alt=True)

    # Right: Inlet conditions
    R = _sec(ws, R, 4, "INLET CONDITIONS", s)
    for i, (lbl, val, fmt) in enumerate([
        ("Inlet pressure (bara)", round(inlet.get("P_in_bara", 0), 3), "0.000"),
        ("Temperature (°C)",      round(inlet.get("T_C", 0), 1),       "0.0"),
        ("Fluid",                 r0.fluid        if r0 else "",        None),
        ("Composition",           r0.composition  if r0 else "",        None),
        ("Flow (kg/h)",           round(r0.flow_kgh,  1) if r0 else 0, "0.0"),
        ("Flow (m³/h) in-situ",  round(r0.flow_m3h, 1) if r0 else 0,  "0.0"),
        ("Correlation",           solver_meta.get("correlation", ""),   None),
        ("Voidage method",        solver_meta.get("voidage", ""),       None),
    ]):
        R = _kv(ws, R, 4, lbl, val, s, alt=bool(i % 2), fmt=fmt)
    R += 1
    R = _sec(ws, R, 4, "HYDRAULIC SUMMARY", s)
    for i, (lbl, val, fmt) in enumerate([
        ("Outlet pressure (bara)",          round(result.outlet_P_bara,     3), "0.000"),
        ("Total ΔP (kPa)",                  round(result.total_dp_kpa,      2), "0.00"),
        ("Total ΔP (bar)",                  round(result.total_dp_kpa/100,  4), "0.0000"),
        ("  ΔP friction + fittings (kPa)",  round(result.total_dp_fric_kpa, 2), "0.00"),
        ("     of which fittings (kPa)",    round(fit_dp,                   2), "0.00"),
        ("  ΔP gravity / equipment (kPa)",  round(result.total_dp_grav_kpa, 2), "0.00"),
        ("Total pipe length (m)",           round(total_L, 1),                  "0.0"),
        ("Max velocity (m/s)",              round(max_v,   2),                  "0.00"),
        ("Max V/V_e",                       round(max_ve,  3),                  "0.000"),
    ]):
        R = _kv(ws, R, 4, lbl, val, s, alt=bool(i % 2), fmt=fmt)

    # Warnings row (right block, below summary)
    R += 1
    R = _sec(ws, R, 4, "WARNINGS", s)
    if result.warnings:
        for w in result.warnings:
            _c(ws, R, 4, w, font=s["warn"], bdr=s["bdr"], align=s["wrap"])
            ws.merge_cells(start_row=R, start_column=4, end_row=R, end_column=5)
            ws.row_dimensions[R].height = 28
            R += 1
    else:
        _c(ws, R, 4, "No warnings.", font=s["ok"], bdr=s["bdr"])
        ws.merge_cells(start_row=R, start_column=4, end_row=R, end_column=5)
        R += 1

    # ── Segment table — starts one row below whichever block is taller ────────
    tbl_start = max(L, R) + 2

    ws.freeze_panes = ws.cell(tbl_start + 1, 2)

    records = [row.to_dict() for row in result.rows]
    props   = [c for c in COLUMNS if c != "Element"]

    # Segment header columns (starting at col B so col A holds property labels)
    seg_col_start = 2
    corner = _c(ws, tbl_start, 1, "Property \\ Segment →",
                font=s["tbl_hdr"], fill=s["tbl_fill"], bdr=s["bdr"], align=s["center"])
    ws.column_dimensions["A"].width = 28
    for j, rec in enumerate(records):
        col = seg_col_start + j
        cl  = openpyxl.utils.get_column_letter(col)
        c = _c(ws, tbl_start, col, f"#{j+1}  {rec['Element']}",
               font=s["tbl_hdr"], fill=s["tbl_fill"], bdr=s["bdr"], align=s["center"])
        ws.column_dimensions[cl].width = 18

    _NUM_PROPS = {
        "ID (mm)", "L (m)", "L_eff (m)", "Δz (m)",
        "Flow (kg/h)", "Flow (m³/h)",
        "P_in (bara)", "P_out (bara)",
        "ΔP_fric (kPa)", "ΔP_fit (kPa)", "ΔP_grav (kPa)", "ΔP (kPa)",
        "V (m/s)", "V/V_e",
    }
    for i, prop in enumerate(props):
        row = tbl_start + 1 + i
        alt  = (i % 2 == 0)
        fill = s["alt_fill"] if alt else None
        lc = _c(ws, row, 1, prop, font=s["lbl"], fill=fill, bdr=s["bdr"])
        for j, rec in enumerate(records):
            val = rec[prop]
            col = seg_col_start + j
            c = _c(ws, row, col, val, fill=fill, bdr=s["bdr"])
            c.font = s["val"]
            if prop in _NUM_PROPS and isinstance(val, float):
                c.number_format = "0.00"
            if prop == "V/V_e" and isinstance(val, (int, float)):
                if val > 1.0:
                    c.fill = PatternFill("solid", fgColor="FEE2E2")
                    c.font = Font(bold=True, size=9, color="B91C1C")
                elif val > 0.8:
                    c.fill = PatternFill("solid", fgColor="FEF3C7")

    ws.row_dimensions[tbl_start].height = 20

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def render(result, line: dict, meta: dict, solver_meta: dict) -> None:
    tag  = (line or {}).get("tag") or "line"
    slug = "".join(ch for ch in tag if ch.isalnum() or ch in "._-") or "line"

    if st.button("📊  Generate Excel", key="qp_gen_xl", width="stretch"):
        st.session_state["qp_xl_bytes"] = build_xlsx(
            result, line, meta, solver_meta).getvalue()
    data = st.session_state.get("qp_xl_bytes")
    if data:
        st.download_button(
            "⬇  Download .xlsx", data=data,
            file_name=f"{slug}_hydraulic.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch", key="qp_dl_xl")
